#This is a file for printing by reflecting the entered value in Excel.

import os
import tkinter
import openpyxl #pip install openpyxl
from num2words import num2words

#Date list -> str
def get_date(date_list):
    if date_list[1][0]=='0':
        date_list[1]=' '+date_list[1][1]
    if date_list[2][0]=='0':
        date_list[2]=' '+date_list[2][1]
    return date_list[0]+'년   '+date_list[1]+'월   '+date_list[2]+'일'

def create_file():
    date = date_entry.get()
    address = address_entry.get()
    price = price_entry.get()

    date_list = [date[:4],date[4:6],date[6:]]
    sum_price = '합계금액:  ' + num2words(int(price), lang="ko")+'원정(\\'+format(int(price), ',')+')'
    surtax = int(price)//11
    #print(date_list)
    #print(sum_price, surtax)

    excelFile = openpyxl.load_workbook('./견적서.xlsx')

    sheet_list = excelFile.sheetnames
    sheet = excelFile[sheet_list[0]]

    #Modify Date
    sheet.cell(row=11, column=1).value = get_date(date_list)

    #Modify Address
    sheet.cell(row=13, column=1).value = address

    #Modify Sum
    sheet.cell(row=19, column=1).value = sum_price

    #Modify Unit
    sheet.cell(row=22, column=4).value = int(price)-surtax

    #Modify Tax
    sheet.cell(row=22, column=6).value = surtax

    #Save File
    excelFile.save(filename= './견적서_결과.xlsx')

    #Print Excel
    os.startfile(os.getcwd()+'\\견적서_결과.xlsx', operation='print')
    
    #Program Shut Down
    window.destroy()
    

window = tkinter.Tk()

window.title("견적서 생성기")
window.geometry("400x200+100+100")
window.resizable(False, False)

date_label = tkinter.Label(window, text="날짜 입력: ", font=('맑은 고딕',16,'bold'))
date_label.grid(row=0,column=0)
date_entry = tkinter.Entry(font=('맑은 고딕',16,'bold'),bg='white',width=10)
date_entry.grid(row=0,column=1,sticky="w")

address_label = tkinter.Label(window, text="주소 입력: ", font=('맑은 고딕',16,'bold'))
address_label.grid(row=1,column=0)
address_entry = tkinter.Entry(font=('맑은 고딕',10,'bold'),bg='white',width=30)
address_entry.grid(row=1,column=1)

price_label = tkinter.Label(window, text="가격 입력: ", font=('맑은 고딕',16,'bold'))
price_label.grid(row=2,column=0)
price_entry = tkinter.Entry(font=('맑은 고딕',16,'bold'),bg='white',width=8)
price_entry.grid(row=2,column=1,sticky="w")

create_button = tkinter.Button(window, text="견적서 생성",command=create_file , font=('맑은 고딕',16,'bold'), bg = "yellow")
create_button.grid(row=3,column=0)

window.mainloop()
